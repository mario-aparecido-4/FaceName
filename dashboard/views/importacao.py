# dashboard_app/views.py
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
from ..models import Aluno, Boletim, Turma, Curso, Serie, Turno, Disciplina, AlunoTurma
from decimal import Decimal, InvalidOperation
import openpyxl
import xlrd
import re
import json
import unicodedata

# --- Funções Auxiliares ---

def _get_cell_value(sheet, row, col):
    """Obtém valor da célula compatível com openpyxl e xlrd."""
    if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
        # openpyxl usa base 1 para indexação
        return sheet.cell(row=row + 1, column=col + 1).value
    else: # xlrd usa base 0
        try:
            return sheet.cell_value(row, col)
        except IndexError:
            return None

def _normalize_str(s):
    """Normaliza strings para comparação (remove acentos e espaços)."""
    if not s:
        return ""
    s = str(s).upper().strip()
    return "".join(
        c for c in unicodedata.normalize('NFD', s)
        if unicodedata.category(c) != 'Mn'
    )

# --- View Principal ---
@login_required
def importar_turma(request):
    contexto = {}
    
    # Recupera dados da sessão se houver preview anterior
    if 'dados_importacao' in request.session:
        try:
            dados_preview = json.loads(request.session['dados_importacao'])
            contexto.update(dados_preview['header'])
            contexto['alunos_dados'] = dados_preview['alunos']
        except (json.JSONDecodeError, KeyError):
            del request.session['dados_importacao']

    if request.method == 'POST':
        try:
            # --- CENÁRIO 1: CONFIRMAR IMPORTAÇÃO (Salvar no Banco) ---
            if 'confirmar_importacao' in request.POST:
                with transaction.atomic():
                    dados_importacao = json.loads(request.session.get('dados_importacao', '{}'))
                    periodo_importacao = request.POST.get('periodo_importacao')

                    if not dados_importacao or not periodo_importacao:
                        raise ValueError("Sessão expirada ou dados não encontrados.")

                    # Define quais campos atualizar baseado na escolha do usuário
                    mapeamento_periodo = {
                        'ate_b1': ['bimestre1', 'faltas', 'faltaspercent'],
                        'ate_b2': ['bimestre1', 'bimestre2', 'faltas', 'faltaspercent'],
                        'ate_r1': ['bimestre1', 'bimestre2', 'recusem1', 'faltas', 'faltaspercent'],
                        'ate_b3': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'faltas', 'faltaspercent'],
                        'ate_b4': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'faltas', 'faltaspercent'],
                        'ate_r2': ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'recusem2', 'faltas', 'faltaspercent'],
                        'final':  ['bimestre1', 'bimestre2', 'recusem1', 'bimestre3', 'bimestre4', 'recusem2', 'recfinal', 'final', 'faltas', 'faltaspercent']
                    }
                    campos_para_atualizar = mapeamento_periodo.get(periodo_importacao, [])
                    header = dados_importacao['header']
                    alunos_dados = dados_importacao['alunos']
                    
                    # Busca objetos relacionados
                    curso_obj = Curso.objects.get(descricao__icontains=header['curso'])
                    serie_obj = Serie.objects.get(descricao__startswith=header['serie'])
                    turno_obj = Turno.objects.get(descricao__iexact=header['turno'])
                    
                    # Cria ou Recupera a Turma
                    try:
                        turma_obj = Turma.objects.get(id=header['turma_id'], ano=header['ano'])
                    except Turma.DoesNotExist:
                        turma_obj = Turma.objects.create(
                            id=header['turma_id'], ano=header['ano'], 
                            descricao=f"{header['turma_id']}",
                            curso=curso_obj, serie=serie_obj, turno=turno_obj
                        )

                    # Lógica de Progressão (Vincula turma anterior se for 2º ou 3º ano)
                    if int(serie_obj.id) in [2, 3]:
                        ano_anterior = int(header['ano']) - 1
                        # Lógica para adivinhar ID da turma anterior: Ex: 512 -> 511
                        # Ajuste conforme sua regra de negócio para ID
                        try:
                            cod_serie_anterior = int(header['turma_id'][-1]) - 1
                            id_turma_anterior = f"{header['turma_id'][:-1]}{cod_serie_anterior}"
                            turma_anterior = Turma.objects.filter(id=id_turma_anterior, ano=ano_anterior).first()
                            if turma_anterior:
                                turma_obj.turma_id = turma_anterior.id # Campo de auto-relacionamento
                                turma_obj.turma_ano = turma_anterior.ano
                                turma_obj.save()
                        except ValueError:
                            pass # Se ID não for numérico no final, ignora
                    
                    disciplinas_db = { _normalize_str(d.descricao): d for d in Disciplina.objects.all() }

                    relatorio = {'alunos_criados': 0, 'alunos_atualizados': 0, 'boletins_atualizados': 0, 'disciplinas_nao_encontradas': set()}
                    
                    # Processa Alunos e Boletins
                    for aluno_data in alunos_dados:
                        aluno_obj, criado_aluno = Aluno.objects.get_or_create(
                            matricula=aluno_data['matricula'], 
                            defaults={'nome': aluno_data['nome']}
                        )
                        if criado_aluno: relatorio['alunos_criados'] += 1
                        else: relatorio['alunos_atualizados'] += 1

                        AlunoTurma.objects.get_or_create(aluno_matricula=aluno_obj, turma_id=turma_obj.id, turma_ano=turma_obj.ano)
                        
                        for boletim_data in aluno_data['boletins']:
                            nome_disciplina_planilha = boletim_data['disciplina']
                            nome_normalizado = _normalize_str(nome_disciplina_planilha)
                            disciplina_obj = disciplinas_db.get(nome_normalizado)

                            if disciplina_obj:
                                identificadores = {
                                    'aluno_matricula': aluno_obj,
                                    'disciplina': disciplina_obj,
                                    'turma_id': turma_obj.id,
                                    'turma_ano': turma_obj.ano
                                }
                                
                                valores_para_atualizar = {'status': aluno_data.get('situacao')}
                                
                                for campo in campos_para_atualizar:
                                    valor = boletim_data.get(campo)
                                    if valor is not None and str(valor).strip() != '':
                                        try:
                                            if campo in ['faltas', 'faltaspercent']:
                                                valores_para_atualizar[campo] = int(float(str(valor).replace(',', '.')))
                                            else:
                                                valores_para_atualizar[campo] = Decimal(str(valor).replace(',', '.'))
                                        except (InvalidOperation, ValueError, TypeError):
                                            continue
                                
                                # Atualiza ou cria Boletim
                                boletim_obj = Boletim.objects.filter(**identificadores).first()
                                if boletim_obj:
                                    Boletim.objects.filter(**identificadores).update(**valores_para_atualizar)
                                else:
                                    Boletim.objects.create(**identificadores, **valores_para_atualizar)
                                
                                relatorio['boletins_atualizados'] += 1
                            else:
                                relatorio['disciplinas_nao_encontradas'].add(nome_disciplina_planilha)
                    
                del request.session['dados_importacao']
                messages.success(request, "Importação concluída com sucesso!")
                
                if relatorio['disciplinas_nao_encontradas']:
                    lista_disc = sorted(list(relatorio['disciplinas_nao_encontradas']))
                    messages.warning(request, f"Disciplinas ignoradas (não encontradas no banco): {', '.join(lista_disc)}")
                    relatorio['disciplinas_nao_encontradas'] = lista_disc

                contexto = {'relatorio': relatorio}

            # --- CENÁRIO 2: UPLOAD DA PLANILHA (Preview) ---
            elif 'planilha' in request.FILES:
                arquivo_excel = request.FILES['planilha']
                if not (arquivo_excel.name.endswith('.xlsx') or arquivo_excel.name.endswith('.xls')):
                    raise ValueError("Formato inválido. Use .xlsx ou .xls.")
                
                if arquivo_excel.name.endswith('.xlsx'):
                    workbook = openpyxl.load_workbook(arquivo_excel, data_only=True)
                    sheet = workbook.active
                else:
                    workbook = xlrd.open_workbook(file_contents=arquivo_excel.read())
                    sheet = workbook.sheet_by_index(0)

                # Extração de Cabeçalho
                ano_entrada_bruto = str(_get_cell_value(sheet, 3, 15) or '') # P4
                serie_bruta = str(_get_cell_value(sheet, 4, 15) or '')       # P5
                
                # Tratamento de erro simples se células estiverem vazias
                if not ano_entrada_bruto or not serie_bruta:
                     raise ValueError("Não foi possível ler Ano ou Série da planilha (Células P4/P5).")

                ano_entrada = int(str(ano_entrada_bruto).split('.')[0])
                match_serie = re.search(r'(\d)', serie_bruta)
                serie_digito = match_serie.group(1) if match_serie else '1'
                serie_int = int(serie_digito)
                ano_corrente = ano_entrada + (serie_int - 1)

                curso_bruto = _get_cell_value(sheet, 2, 1) or ''
                turno_bruto = (_get_cell_value(sheet, 3, 1) or '').upper()
                
                match_curso = re.search(r'EM\s(.*?)\s\(', curso_bruto)
                curso_extraido = match_curso.group(1).strip() if match_curso else "Não encontrado"
                
                serie_extraida = f"{serie_digito}º"
                turno_extraido = "MATUTINO" if "MATUTINO" in turno_bruto else "VESPERTINO" if "VESPERTINO" in turno_bruto else "NOTURNO" if "NOTURNO" in turno_bruto else "Não encontrado"
                
                contexto.update({'curso': curso_extraido, 'ano': ano_corrente, 'serie': serie_extraida, 'turno': turno_extraido})
                
                # Mapeamento de códigos (Ajuste conforme seus IDs reais no banco)
                mapa_cod_curso = {"INFORMÁTICA": "5", "ELETROTÉCNICA": "4", "EDIFICAÇÕES": "2", "SEGURANÇA DO TRABALHO": "S"}
                mapa_cod_turno = {"MATUTINO": "1", "VESPERTINO": "2", "NOTURNO": "3"}
                
                cod_curso = mapa_cod_curso.get(curso_extraido.upper(), "?")
                cod_turno = mapa_cod_turno.get(turno_extraido, "?")
                cod_serie = serie_digito
                
                turma_id_gerado = f"{cod_curso}{cod_turno}{cod_serie}"
                contexto['turma_id'] = turma_id_gerado

                # Busca de Disciplinas (Header da tabela)
                disciplinas = []
                max_cols = sheet.ncols if hasattr(sheet, 'ncols') else sheet.max_column
                col_atual = 3 # Começa na coluna D (índice 3)
                
                # Loop para encontrar colunas de disciplinas
                while col_atual < max_cols:
                    nome_disciplina = _get_cell_value(sheet, 6, col_atual) # Linha 7 (index 6)
                    # Verifica se chegou na coluna de Situação ou se acabou
                    header_check = _get_cell_value(sheet, 7, col_atual)
                    if header_check and "SITUAÇÃO" in str(header_check).upper():
                        contexto['coluna_situacao'] = col_atual
                        break
                    
                    if nome_disciplina and str(nome_disciplina).strip():
                        disciplinas.append({'nome': str(nome_disciplina).strip(), 'col_inicio': col_atual})
                    
                    col_atual += 11 # Pula o bloco de notas da disciplina
                
                # Extração de Alunos
                alunos_dados = []
                max_rows = sheet.nrows if hasattr(sheet, 'nrows') else sheet.max_row
                
                for row in range(8, max_rows): # Começa na linha 9 (index 8)
                    matricula = _get_cell_value(sheet, row, 1)
                    if not matricula: continue # Pula linha vazia
                    
                    nome = str(_get_cell_value(sheet, row, 2)).strip()
                    aluno_atual = {
                        'matricula': str(int(matricula)) if isinstance(matricula, (int, float)) else str(matricula), 
                        'nome': nome, 
                        'boletins': []
                    }
                    
                    if 'coluna_situacao' in contexto:
                        aluno_atual['situacao'] = _get_cell_value(sheet, row, contexto['coluna_situacao'])
                    
                    for disc in disciplinas:
                        c = disc['col_inicio']
                        # Mapeamento relativo das colunas da disciplina
                        notas = {
                            'disciplina': disc['nome'],
                            'bimestre1': _get_cell_value(sheet, row, c),
                            'bimestre2': _get_cell_value(sheet, row, c+1),
                            'recusem1':  _get_cell_value(sheet, row, c+2),
                            'bimestre3': _get_cell_value(sheet, row, c+3),
                            'bimestre4': _get_cell_value(sheet, row, c+4),
                            'recusem2':  _get_cell_value(sheet, row, c+5),
                            'recfinal':  _get_cell_value(sheet, row, c+7),
                            'final':     _get_cell_value(sheet, row, c+8),
                            'faltas':    _get_cell_value(sheet, row, c+9),
                            'faltaspercent': _get_cell_value(sheet, row, c+10)
                        }
                        aluno_atual['boletins'].append(notas)
                    alunos_dados.append(aluno_atual)
                
                contexto['alunos_dados'] = alunos_dados
                
                # Salva na sessão para o próximo passo (POST de confirmação)
                dados_para_salvar = {'header': {k: v for k, v in contexto.items() if k != 'alunos_dados'}, 'alunos': alunos_dados}
                request.session['dados_importacao'] = json.dumps(dados_para_salvar)

        except Exception as e:
            messages.error(request, f"Erro ao processar: {str(e)}")
            # Limpa sessão em caso de erro
            if 'dados_importacao' in request.session: del request.session['dados_importacao']
    
    return render(request, 'importacao/importar_turma.html', contexto)